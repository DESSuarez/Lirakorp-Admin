import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

df = pd.read_excel('/Users/diego/Downloads/Relacion RENTAS 2026.xlsx', header=None)

# Zone sections detected by inspection:
# Row 0: "BODEGAS AV. WASHINGTON" -> rows 3-31
# Row 35: "ACK CIMENTACIONES SA DE CV BODEGAS VALLARTA" -> rows 36-43
# Row 46: "EDIFICIO FLUVIAL (ACK)" -> rows 47-52
# Row 55: "PLAZA ALTURES (ACK)" -> rows 56-72
# Row 78: "CANCUN WOLFTOWER(ZAIKE)" -> rows 79-82
# Row 84: "OFICINAS NIDO (ZAIKE)" -> rows 85-86
# Row 88: "MASTER PLAZA(ZAIKE)" -> rows 89-91
# Row 95: "RAUL LIRA DIRECTO" -> rows 97-100

zones = [
    ("Bodegas Av. Washington", 3, 31),
    ("ACK Bodegas Vallarta", 36, 43),
    ("Edificio Fluvial (ACK)", 47, 52),
    ("Plaza Altures (ACK)", 56, 72),
    ("Cancun WolfTower (Zaike)", 79, 82),
    ("Oficinas Nido (Zaike)", 85, 86),
    ("Master Plaza (Zaike)", 89, 91),
    ("Raul Lira Directo", 97, 100),
]

properties = []

def clean_text(v):
    if pd.isna(v) or v is None:
        return ''
    s = str(v).strip()
    s = s.replace('\xa0', ' ').strip()
    return s

def parse_date_text(v):
    """Parse Spanish date text to dd/mm/yyyy"""
    if pd.isna(v) or not v:
        return ''
    s = str(v).strip()
    if not s:
        return ''

    months = {
        'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
        'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
        'septiembre': '09', 'setiembre': '09', 'octubre': '10',
        'noviembre': '11', 'diciembre': '12'
    }

    s_lower = s.lower().strip()

    # Skip non-date values
    if 'sin contrato' in s_lower or 'autorizo' in s_lower:
        return ''

    # Pattern: "DD de MONTH YYYY" or "DD de MONTH de YYYY"
    for month_name, month_num in months.items():
        pattern = rf'(\d{{1,2}})\s*de\s*{month_name}\s*(?:de\s*)?(\d{{4}})'
        match = re.search(pattern, s_lower)
        if match:
            day = match.group(1).zfill(2)
            year = match.group(2)
            return f'{day}/{month_num}/{year}'

    # Pattern: "Del DD de MONTH YYYY"
    for month_name, month_num in months.items():
        pattern = rf'del?\s*(\d{{1,2}})\s*de\s*{month_name}\s*(?:de\s*)?(\d{{4}})'
        match = re.search(pattern, s_lower)
        if match:
            day = match.group(1).zfill(2)
            year = match.group(2)
            return f'{day}/{month_num}/{year}'

    # Pattern: "Al DD de MONTH YYYY"
    for month_name, month_num in months.items():
        pattern = rf'al?\s*(\d{{1,2}})\s*de\s*{month_name}\s*(?:de\s*)?(\d{{4}})'
        match = re.search(pattern, s_lower)
        if match:
            day = match.group(1).zfill(2)
            year = match.group(2)
            return f'{day}/{month_num}/{year}'

    return ''

def get_rent(row):
    """Get rent from column I (Renta NETA) - col index 8"""
    v = row.iloc[8] if len(row) > 8 else None
    if pd.isna(v):
        return 0
    try:
        return float(v)
    except:
        return 0

def get_total_rent(row):
    """Get total rent from column J (Renta total) - col index 9"""
    v = row.iloc[9] if len(row) > 9 else None
    if pd.isna(v):
        return 0
    try:
        return float(v)
    except:
        return 0

for zone_name, start_row, end_row in zones:
    for i in range(start_row, end_row + 1):
        if i >= len(df):
            continue
        row = df.iloc[i]

        number = clean_text(row.iloc[2])  # Col C
        tenant = clean_text(row.iloc[3])  # Col D
        m2_raw = row.iloc[4] if len(row) > 4 else None  # Col E
        giro = clean_text(row.iloc[5]) if len(row) > 5 else ''  # Col F
        start_date = clean_text(row.iloc[6]) if len(row) > 6 else ''  # Col G
        end_date = clean_text(row.iloc[7]) if len(row) > 7 else ''  # Col H
        rent_neta = get_rent(row)
        rent_total = get_total_rent(row)
        email = clean_text(row.iloc[11]) if len(row) > 11 else ''  # Col L
        email2 = clean_text(row.iloc[12]) if len(row) > 12 else ''  # Col M

        # Skip empty rows, summary rows, headers
        if not number and not tenant:
            continue
        if tenant.upper() in ['', 'NOMBRE']:
            continue

        # Use number as property ID, clean commas
        prop_number = number.replace(',', '').strip()
        if not prop_number and tenant:
            # Some rows use tenant name area as identifier (like "Local la peñita")
            prop_number = number if number else tenant[:20]

        # Determine property type
        ptype = 'bodega'
        if 'local' in prop_number.lower() or 'local' in number.lower():
            ptype = 'local'
        elif 'oficina' in zone_name.lower() or 'nido' in zone_name.lower():
            ptype = 'local'
        elif 'depa' in zone_name.lower() or 'wolf' in zone_name.lower():
            ptype = 'departamento'
        elif 'plaza' in zone_name.lower():
            ptype = 'local'
        elif 'casa' in prop_number.lower():
            ptype = 'casa'

        # Parse m2
        m2 = 0
        if pd.notna(m2_raw):
            try:
                m2 = float(m2_raw)
            except:
                pass

        # Use rent_total if available, otherwise rent_neta
        rent = rent_total if rent_total > 0 else rent_neta

        # Parse dates
        start_parsed = parse_date_text(start_date)
        end_parsed = parse_date_text(end_date)

        # Clean email - remove notes and extra text
        clean_email = ''
        if email and '@' in email:
            # Extract first email from string
            email_match = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', email)
            if email_match:
                clean_email = email_match.group(0)

        clean_email2 = ''
        if email2 and '@' in email2:
            email_match2 = re.search(r'[\w.+-]+@[\w-]+\.[\w.-]+', email2)
            if email_match2:
                clean_email2 = email_match2.group(0)

        # Create zone-prefixed number for uniqueness
        zone_prefix = {
            "Bodegas Av. Washington": "BW",
            "ACK Bodegas Vallarta": "ABV",
            "Edificio Fluvial (ACK)": "EF",
            "Plaza Altures (ACK)": "PA",
            "Cancun WolfTower (Zaike)": "CW",
            "Oficinas Nido (Zaike)": "ON",
            "Master Plaza (Zaike)": "MP",
            "Raul Lira Directo": "RL",
        }
        prefix = zone_prefix.get(zone_name, "XX")
        unique_number = f"{prefix}-{prop_number}" if prop_number else f"{prefix}-{i}"

        # Property name
        prop_name = f"{zone_name} {number}" if number else f"{zone_name} {tenant[:30]}"

        # Skip VACIO/VACIA/PRESTADO entries (no active tenant)
        is_vacant = tenant.upper() in ['VACIA', 'VACIO', 'VACIOS'] or 'PRESTADO' in tenant.upper()

        properties.append({
            'numero': unique_number,
            'nombre': prop_name.strip(),
            'm2': m2,
            'tipo': ptype,
            'zona': zone_name,
            'direccion': '',
            'inquilino': '' if is_vacant else tenant,
            'email': '' if is_vacant else clean_email,
            'telefono': '',
            'inicio': '' if is_vacant else start_parsed,
            'fin': '' if is_vacant else end_parsed,
            'renta': rent,
            'giro': giro,
            'email2': '' if is_vacant else clean_email2,
            'estado': 'VACIO' if is_vacant else 'OCUPADO',
        })

# Create output Excel
wb = Workbook()
ws = wb.active
ws.title = 'Importar'

headers = ['Numero', 'Nombre', 'M2', 'Tipo', 'Zona', 'Direccion', 'Inquilino', 'Email', 'Telefono', 'Inicio', 'Fin', 'Renta']
header_fill = PatternFill('solid', fgColor='1E40AF')
header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

widths = [14, 35, 8, 14, 28, 20, 40, 35, 16, 12, 12, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+i)].width = w

data_font = Font(size=10, name='Arial')
for idx, p in enumerate(properties):
    row = idx + 2
    vals = [
        p['numero'], p['nombre'], p['m2'], p['tipo'], p['zona'],
        p['direccion'], p['inquilino'], p['email'], p['telefono'],
        p['inicio'], p['fin'], p['renta']
    ]
    for col, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = data_font
        cell.border = thin_border
        if col == 12:  # Renta
            cell.number_format = '#,##0.00'

# Add second sheet with full data including extra emails and giro
ws2 = wb.create_sheet('Datos Completos')
headers2 = ['Numero', 'Nombre', 'M2', 'Tipo', 'Zona', 'Inquilino', 'Email Principal', 'Email Secundario', 'Giro', 'Inicio', 'Fin', 'Renta', 'Estado']
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for idx, p in enumerate(properties):
    row = idx + 2
    vals = [
        p['numero'], p['nombre'], p['m2'], p['tipo'], p['zona'],
        p['inquilino'], p['email'], p['email2'], p['giro'],
        p['inicio'], p['fin'], p['renta'], p['estado']
    ]
    for col, v in enumerate(vals, 1):
        cell = ws2.cell(row=row, column=col, value=v)
        cell.font = data_font
        cell.border = thin_border

out = '/Users/diego/Downloads/LIRAKORP_Importar_Rentas_2026.xlsx'
wb.save(out)

print(f'Total propiedades extraidas: {len(properties)}')
print(f'Archivo guardado en: {out}')
print()
# Summary by zone
from collections import Counter
zone_counts = Counter(p['zona'] for p in properties)
for z, c in zone_counts.items():
    occupied = sum(1 for p in properties if p['zona'] == z and p['inquilino'])
    vacant = sum(1 for p in properties if p['zona'] == z and not p['inquilino'])
    print(f'  {z}: {c} propiedades ({occupied} ocupadas, {vacant} vacias)')
print()
# Print all for verification
for p in properties:
    tenant = p['inquilino'][:35] if p['inquilino'] else 'VACIO'
    print(f"  {p['numero']:15s} | {p['nombre'][:30]:30s} | {tenant:35s} | {p['inicio']:12s} | {p['fin']:12s} | ${p['renta']:>10,.2f}")
